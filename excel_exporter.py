import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional, Union, Any
import os


class ExcelExporter:
    MAX_SHEET_NAME_LENGTH = 31
    INVALID_SHEET_CHARS = ["\\", "/", "*", "?", ":", "[", "]"]

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.sheets: List[Dict[str, Any]] = []
        self._used_names: set = set()

    def _sanitize_sheet_name(self, name: str) -> str:
        if not name or not isinstance(name, str):
            name = "Sheet"

        for char in self.INVALID_SHEET_CHARS:
            name = name.replace(char, "_")

        if len(name) > self.MAX_SHEET_NAME_LENGTH:
            name = name[: self.MAX_SHEET_NAME_LENGTH]

        original_name = name
        counter = 1
        while name in self._used_names:
            suffix = f"_{counter}"
            if len(original_name) + len(suffix) > self.MAX_SHEET_NAME_LENGTH:
                name = original_name[: self.MAX_SHEET_NAME_LENGTH - len(suffix)] + suffix
            else:
                name = original_name + suffix
            counter += 1

        self._used_names.add(name)
        return name

    def add_sheet(self, data: Union[pd.DataFrame, List[Dict], Dict], sheet_name: str, **kwargs) -> None:
        if isinstance(data, pd.DataFrame):
            df = data
        elif isinstance(data, list) and all(isinstance(item, dict) for item in data):
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            df = pd.DataFrame([data])
        else:
            raise ValueError("data must be a DataFrame, list of dicts, or dict")

        safe_name = self._sanitize_sheet_name(sheet_name)

        sheet_config = {
            "name": safe_name,
            "data": df,
            "header_style": kwargs.get("header_style", True),
            "auto_width": kwargs.get("auto_width", True),
            "freeze_panes": kwargs.get("freeze_panes", "A2"),
        }
        self.sheets.append(sheet_config)

    def export(self) -> str:
        if not self.sheets:
            raise ValueError("No sheets to export. Add sheets first using add_sheet().")

        with pd.ExcelWriter(self.file_path, engine="openpyxl") as writer:
            for sheet in self.sheets:
                self._write_sheet(writer, sheet)

        return os.path.abspath(self.file_path)

    def _write_sheet(self, writer: pd.ExcelWriter, sheet: Dict[str, Any]) -> None:
        df = sheet["data"]
        sheet_name = sheet["name"]

        df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        if sheet["header_style"]:
            self._apply_header_style(worksheet, df)

        if sheet["auto_width"]:
            self._adjust_column_width(worksheet, df)

        if sheet["freeze_panes"]:
            worksheet.freeze_panes = sheet["freeze_panes"]

    def _apply_header_style(self, worksheet, df: pd.DataFrame) -> None:
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_num, _ in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _adjust_column_width(self, worksheet, df: pd.DataFrame) -> None:
        for col_num, column in enumerate(df.columns, 1):
            header_length = len(str(column))
            max_data_length = df[column].astype(str).str.len().max() if len(df) > 0 else 0
            max_length = max(header_length, max_data_length)
            adjusted_width = min(max(max_length + 2, 8), 50)
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(
    data_sheets: List[Dict[str, Any]],
    file_path: str,
) -> str:
    exporter = ExcelExporter(file_path)
    for sheet in data_sheets:
        exporter.add_sheet(
            data=sheet["data"],
            sheet_name=sheet["name"],
            header_style=sheet.get("header_style", True),
            auto_width=sheet.get("auto_width", True),
            freeze_panes=sheet.get("freeze_panes", "A2"),
        )
    return exporter.export()
